#!/usr/bin/env python3
from __future__ import annotations

"""Build aligned Abaqus displacement datasets for the thesis figures.

本脚本读取 S1 至 S6 工况的 Abaqus 位移导出 Excel 表，按工况、测线、
原始帧号和采样点整理成长表。随后计算中面位移、壁厚变化近似量和
帧级位移能量，自动识别每个工况的主响应区间，并把不同工况映射到
统一的归一化切削进程 `xi`。输出结果用于论文第 3 章中面位移响应、
时间对齐和参数影响分析。

This script reads Abaqus displacement-export workbooks for cases S1-S6 and
organizes them into long tables indexed by case, measurement line, raw frame,
and sample point. It computes mid-surface displacement, approximate thickness
change, frame-level displacement energy, the main response window, and a
normalized cutting-progress coordinate `xi` for cross-case comparison.
"""

import argparse
import json
import os
from pathlib import Path

# Put Matplotlib's font cache inside the project output folder. This avoids
# writing cache files to a user-specific home directory in restricted runtimes.
# 将 Matplotlib 字体缓存放进项目 output，避免受限环境写入用户主目录。
os.environ.setdefault("MPLCONFIGDIR", str((Path(__file__).resolve().parents[2] / "output" / ".mplconfig")))

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from scipy.interpolate import PchipInterpolator

from common import ROOT, case_settings, iter_displacement_files, line_positions, load_case_config, normalize_case_id


def parse_args() -> argparse.Namespace:
    """Parse pipeline parameters and anonymized default paths.

    `--grid-size` 控制对齐后的统一网格密度；`--smooth-window` 和
    `--threshold` 控制主响应窗口识别。

    `--grid-size` controls the aligned common grid. `--smooth-window` and
    `--threshold` control main-response-window detection.
    """
    parser = argparse.ArgumentParser(description="构建位移对齐数据集并输出质检图表。")
    parser.add_argument(
        "--input",
        default=str(ROOT / "data" / "abaqus" / "displacement_exports"),
        help="位移导出表目录。",
    )
    parser.add_argument(
        "--config",
        default=str(ROOT / "scripts" / "abaqus" / "case_config.yml"),
        help="案例配置文件路径。",
    )
    parser.add_argument(
        "--outdir",
        default=str(ROOT / "output" / "spreadsheet" / "abaqus"),
        help="表格输出目录。",
    )
    parser.add_argument(
        "--figdir",
        default=str(ROOT / "output" / "figures" / "abaqus"),
        help="质检图输出目录。",
    )
    parser.add_argument("--grid-size", type=int, default=None, help="统一 xi 网格点数。")
    parser.add_argument("--smooth-window", type=int, default=None, help="能量序列平滑窗口。")
    parser.add_argument("--threshold", type=float, default=None, help="主响应区间阈值比例。")
    return parser.parse_args()


def configure_matplotlib() -> None:
    """Set fonts and figure defaults for reproducible exports.

    中文图件优先使用宋体兼容字体；负号显示、背景和图例样式统一设置。

    Chinese figures prefer Songti-compatible fonts. Minus signs, backgrounds,
    and legend styles are fixed for reproducible output.
    """
    mpl.rcParams["font.family"] = ["Songti SC", "Arial Unicode MS", "DejaVu Serif"]
    mpl.rcParams["axes.unicode_minus"] = False
    mpl.rcParams["figure.facecolor"] = "white"
    mpl.rcParams["axes.facecolor"] = "white"
    mpl.rcParams["legend.frameon"] = False


def read_displacement_table(xlsx_path: Path) -> pd.DataFrame:
    """Read one Abaqus displacement workbook.

    原始表包含两侧表面位移。脚本计算中面位移 `w_mid_mm` 和两侧位移差
    `delta_t_mm`，分别用于让刀响应和壁厚变化近似分析。

    The raw workbook contains displacement on two opposite surfaces. The script
    computes `w_mid_mm` for tool-deflection response and `delta_t_mm` as an
    approximate thickness-change indicator.
    """
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    frame_df = pd.DataFrame(rows, columns=["x_mm", "u_plus_mm", "x_dup_mm", "u_minus_mm"])
    frame_df["sample_index"] = np.arange(1, len(frame_df) + 1, dtype=int)
    frame_df["w_mid_mm"] = (frame_df["u_plus_mm"] + frame_df["u_minus_mm"]) / 2.0
    frame_df["delta_t_mm"] = frame_df["u_plus_mm"] - frame_df["u_minus_mm"]
    return frame_df[["sample_index", "x_mm", "u_plus_mm", "u_minus_mm", "w_mid_mm", "delta_t_mm"]]


def build_raw_long(input_dir: Path, config: dict[str, object]) -> pd.DataFrame:
    """Stack all workbooks into one long displacement table.

    每行对应一个工况、测线、原始帧和采样点的位移记录。

    Each row represents one displacement record at a case, line, raw frame, and
    sample point.
    """
    line_mm_map = line_positions(config)
    records: list[pd.DataFrame] = []
    for case_id, line_id, frame_raw, xlsx_path in iter_displacement_files(input_dir):
        frame_df = read_displacement_table(xlsx_path)
        frame_df.insert(0, "case", normalize_case_id(case_id))
        frame_df.insert(1, "line_id", line_id)
        frame_df.insert(2, "line_mm", float(line_mm_map[line_id]))
        frame_df.insert(3, "frame_raw", int(frame_raw))
        records.append(frame_df)
    if not records:
        raise RuntimeError(f"未在 {input_dir} 找到任何位移表。")
    raw_long = pd.concat(records, ignore_index=True)
    return raw_long.sort_values(["case", "line_mm", "frame_raw", "sample_index"]).reset_index(drop=True)


def apply_case_path_cropping(raw_long: pd.DataFrame, config: dict[str, object]) -> pd.DataFrame:
    """Crop the middle portion of overlong case paths when configured.

    对于个别总切削路径偏长的算例，可在配置文件中给出 `path_keep_fraction`。
    脚本保留 x 方向中段数据，并重新编号采样点，使后续能量窗口、峰值位移
    和测线统计使用同一有效路径口径。

    For a case with an overlong cutting path, `path_keep_fraction` keeps the
    centred part of the x-coordinate range and rebuilds point indices. Response
    windows, peak displacement, and line metrics then share the same path basis.
    """
    trimmed_groups: list[pd.DataFrame] = []
    for case_id, case_group in raw_long.groupby("case", sort=True):
        case_cfg = case_settings(config, case_id)
        keep_fraction = case_cfg.get("path_keep_fraction")
        if keep_fraction in (None, ""):
            trimmed_groups.append(case_group.copy())
            continue

        keep_fraction = float(keep_fraction)
        if keep_fraction <= 0.0 or keep_fraction > 1.0:
            raise ValueError(f"{case_id} 的 path_keep_fraction 非法: {keep_fraction}")
        if keep_fraction >= 0.999999:
            trimmed_groups.append(case_group.copy())
            continue

        x_min = float(case_group["x_mm"].min())
        x_max = float(case_group["x_mm"].max())
        x_center = 0.5 * (x_min + x_max)
        x_half_span = 0.5 * (x_max - x_min) * keep_fraction
        lower = x_center - x_half_span
        upper = x_center + x_half_span

        trimmed = case_group[case_group["x_mm"].between(lower, upper)].copy()
        trimmed = trimmed.sort_values(["line_mm", "frame_raw", "x_mm", "sample_index"]).reset_index(drop=True)
        if trimmed.empty:
            raise ValueError(f"{case_id} 按路径中段裁剪后无有效位移数据。")

        trimmed["sample_index"] = (
            trimmed.groupby(["case", "line_id", "frame_raw"], sort=False).cumcount() + 1
        ).astype(int)
        trimmed_groups.append(trimmed)

    return (
        pd.concat(trimmed_groups, ignore_index=True)
        .sort_values(["case", "line_mm", "frame_raw", "sample_index"])
        .reset_index(drop=True)
    )


def compute_raw_metrics(raw_long: pd.DataFrame) -> pd.DataFrame:
    """Compute frame-level displacement and thickness-change metrics.

    指标包括中面位移峰值、平均绝对值、均方根值和两侧位移差统计量，
    用于正文表格、热力图和窗口识别。

    Metrics include peak, mean absolute, RMS mid-surface displacement and
    two-surface displacement-difference statistics. They support tables,
    heatmaps, and response-window detection.
    """
    metric_rows: list[dict[str, object]] = []
    group_cols = ["case", "line_id", "line_mm", "frame_raw"]
    for (case_id, line_id, line_mm, frame_raw), group in raw_long.groupby(group_cols, sort=True):
        w_values = group["w_mid_mm"].to_numpy(dtype=float)
        dt_values = group["delta_t_mm"].to_numpy(dtype=float)
        x_values = group["x_mm"].to_numpy(dtype=float)
        w_abs = np.abs(w_values)
        dt_abs = np.abs(dt_values)
        peak_index = int(np.argmax(w_abs))
        metric_rows.append(
            {
                "case": case_id,
                "line_id": line_id,
                "line_mm": float(line_mm),
                "frame_raw": int(frame_raw),
                "point_count": int(len(group)),
                "w_peak_abs_mm": float(w_abs.max()),
                "w_mean_abs_mm": float(w_abs.mean()),
                "w_rms_mm": float(np.sqrt(np.mean(np.square(w_values)))),
                "w_l1_mm": float(w_abs.sum()),
                "w_signed_mean_mm": float(w_values.mean()),
                "delta_t_peak_abs_mm": float(dt_abs.max()),
                "delta_t_mean_abs_mm": float(dt_abs.mean()),
                "delta_t_signed_mean_mm": float(dt_values.mean()),
                "delta_t_l1_mm": float(dt_abs.sum()),
                "x_at_w_peak_mm": float(x_values[peak_index]),
            }
        )
    return pd.DataFrame(metric_rows).sort_values(group_cols).reset_index(drop=True)


def compute_case_energy(raw_long: pd.DataFrame, smooth_window: int) -> pd.DataFrame:
    """Aggregate each frame into a case-level displacement energy.

    `energy_l1_mm` 是该帧所有测线和采样点的中面位移绝对值之和。平滑后
    的能量曲线用于寻找主响应区间。

    `energy_l1_mm` is the sum of absolute mid-surface displacement over all
    lines and sample points in a frame. The smoothed energy curve is used to
    locate the main response window.
    """
    energy_rows: list[dict[str, object]] = []
    for (case_id, frame_raw), group in raw_long.groupby(["case", "frame_raw"], sort=True):
        w_values = group["w_mid_mm"].to_numpy(dtype=float)
        w_abs = np.abs(w_values)
        energy_rows.append(
            {
                "case": case_id,
                "frame_raw": int(frame_raw),
                "energy_l1_mm": float(w_abs.sum()),
                "energy_l2_mm": float(np.sqrt(np.square(w_values).sum())),
                "peak_abs_mm": float(w_abs.max()),
                "mean_abs_mm": float(w_abs.mean()),
            }
        )
    energy_df = pd.DataFrame(energy_rows).sort_values(["case", "frame_raw"]).reset_index(drop=True)
    energy_df["energy_smooth_mm"] = (
        energy_df.groupby("case")["energy_l1_mm"]
        .transform(lambda series: series.rolling(window=smooth_window, center=True, min_periods=1).mean())
    )
    return energy_df


def contiguous_peak_window(case_energy: pd.DataFrame, threshold_ratio: float, case_cfg: dict[str, object]) -> dict[str, object]:
    """Detect the continuous response window around the energy peak.

    以平滑能量峰值为中心，向两侧扩展到低于阈值的位置，得到 `ks_raw`
    和 `ke_raw`。配置中的人工覆盖值优先级更高，用于保留人工校正口径。

    Starting from the smoothed-energy peak, the window expands until values fall
    below the threshold. Configuration overrides take precedence when manual
    correction is needed.
    """
    case_energy = case_energy.sort_values("frame_raw").reset_index(drop=True)
    smooth_values = case_energy["energy_smooth_mm"].to_numpy(dtype=float)
    frame_values = case_energy["frame_raw"].to_numpy(dtype=int)
    peak_pos = int(np.argmax(smooth_values))
    peak_frame = int(frame_values[peak_pos])
    peak_energy = float(smooth_values[peak_pos])
    threshold_value = peak_energy * float(threshold_ratio)
    mask = smooth_values >= threshold_value

    start = peak_pos
    while start > 0 and mask[start - 1]:
        start -= 1
    end = peak_pos
    while end < len(mask) - 1 and mask[end + 1]:
        end += 1

    ks_auto = int(frame_values[start])
    ke_auto = int(frame_values[end])
    ks_override = case_cfg.get("ks_override")
    ke_override = case_cfg.get("ke_override")
    override_used = ks_override is not None or ke_override is not None
    ks_raw = int(ks_override if ks_override is not None else ks_auto)
    ke_raw = int(ke_override if ke_override is not None else ke_auto)

    if ke_raw <= ks_raw:
        raise ValueError(f"{case_energy['case'].iloc[0]} 的 ks/ke 非法: ks={ks_raw}, ke={ke_raw}")

    return {
        "case": case_energy["case"].iloc[0],
        "ks_raw": ks_raw,
        "ke_raw": ke_raw,
        "peak_frame_raw": peak_frame,
        "peak_energy": peak_energy,
        "threshold_ratio": float(threshold_ratio),
        "threshold_value": threshold_value,
        "override_used": bool(override_used),
        "notes": "平滑能量主峰连续区间",
    }


def align_raw_long(raw_long: pd.DataFrame, summary_df: pd.DataFrame) -> pd.DataFrame:
    """Filter raw frames to each case window and compute normalized progress.

    `xi=0` 对应主响应起点，`xi=1` 对应主响应终点。该坐标消除不同工况
    原始帧范围不一致造成的比较偏差。

    `xi=0` marks the start of the main response window and `xi=1` marks the end.
    This coordinate reduces bias caused by unequal raw frame ranges.
    """
    aligned = raw_long.merge(summary_df[["case", "ks_raw", "ke_raw"]], on="case", how="left")
    aligned = aligned[(aligned["frame_raw"] >= aligned["ks_raw"]) & (aligned["frame_raw"] <= aligned["ke_raw"])].copy()
    aligned["xi"] = (aligned["frame_raw"] - aligned["ks_raw"]) / (aligned["ke_raw"] - aligned["ks_raw"])
    return aligned.sort_values(["case", "line_mm", "sample_index", "xi"]).reset_index(drop=True)


def interpolate_to_grid(aligned_raw: pd.DataFrame, grid_size: int) -> pd.DataFrame:
    """Interpolate aligned histories to a shared `xi` grid.

    对同一工况、测线和采样点的时程使用 PCHIP 插值。PCHIP 在保持曲线
    单调特征方面比普通高阶多项式更稳健。

    PCHIP interpolation is applied to each case-line-sample history. PCHIP is
    more stable than high-order polynomial interpolation for preserving local
    shape.
    """
    xi_grid = np.linspace(0.0, 1.0, grid_size)
    interp_rows: list[pd.DataFrame] = []
    group_cols = ["case", "line_id", "line_mm", "sample_index"]
    for (case_id, line_id, line_mm, sample_index), group in aligned_raw.groupby(group_cols, sort=True):
        group = group.sort_values("xi")
        xi_values = group["xi"].to_numpy(dtype=float)
        interp_columns = {}
        for column in ["x_mm", "u_plus_mm", "u_minus_mm", "w_mid_mm", "delta_t_mm"]:
            interpolator = PchipInterpolator(xi_values, group[column].to_numpy(dtype=float))
            interp_columns[column] = interpolator(xi_grid)
        interp_df = pd.DataFrame(interp_columns)
        interp_df.insert(0, "xi", xi_grid)
        interp_df.insert(0, "sample_index", int(sample_index))
        interp_df.insert(0, "line_mm", float(line_mm))
        interp_df.insert(0, "line_id", line_id)
        interp_df.insert(0, "case", case_id)
        interp_rows.append(interp_df)
    return pd.concat(interp_rows, ignore_index=True)


def build_acceptance_checks(raw_metrics: pd.DataFrame) -> dict[str, object]:
    """Compare peak displacement values with retained reference values.

    该检查用于发现输入目录或配置发生变化后是否导致主要数值口径漂移。

    This check helps detect whether input or configuration changes alter the
    retained numerical baseline.
    """
    expected = {
        "S1": (0.4099, 190),
        "S2": (0.1826, 190),
        "S3": (0.6454, 200),
        "S4": (0.1802, 180),
        "S5": (0.7186, 150),
        "S6": (0.2294, 160),
    }
    observed: dict[str, dict[str, object]] = {}
    for case_id, group in raw_metrics.groupby("case"):
        peak_row = group.loc[group["w_peak_abs_mm"].idxmax()]
        observed[case_id] = {
            "peak_w_mm": float(peak_row["w_peak_abs_mm"]),
            "peak_frame_raw": int(peak_row["frame_raw"]),
            "matches_expected": (
                abs(float(peak_row["w_peak_abs_mm"]) - expected[case_id][0]) < 5e-4
                and int(peak_row["frame_raw"]) == expected[case_id][1]
            ),
        }
    return {"expected": expected, "observed": observed}


def plot_qc_energy_windows(case_energy: pd.DataFrame, summary_df: pd.DataFrame, figdir: Path) -> None:
    """Export a quality-control plot for response-window detection.

    图中同时显示原始能量、平滑能量、阈值、起止帧和峰值帧，便于人工复核。

    The plot shows raw energy, smoothed energy, threshold, window boundaries,
    and peak frame for manual review.
    """
    figdir.mkdir(parents=True, exist_ok=True)
    figure, axes = plt.subplots(2, 3, figsize=(14, 8), sharex=False, sharey=False)
    axes = axes.ravel()
    summary_lookup = summary_df.set_index("case").to_dict("index")

    for axis, case_id in zip(axes, sorted(case_energy["case"].unique())):
        group = case_energy[case_energy["case"] == case_id].sort_values("frame_raw")
        summary = summary_lookup[case_id]
        axis.plot(group["frame_raw"], group["energy_l1_mm"], color="#b0b0b0", linewidth=1.2, label="原始能量")
        axis.plot(group["frame_raw"], group["energy_smooth_mm"], color="#1f4e79", linewidth=2.0, label="平滑能量")
        axis.axhline(summary["threshold_value"], color="#c0504d", linestyle="--", linewidth=1.0, label="阈值")
        axis.axvline(summary["ks_raw"], color="#4f81bd", linestyle=":", linewidth=1.0)
        axis.axvline(summary["ke_raw"], color="#4f81bd", linestyle=":", linewidth=1.0)
        axis.scatter(
            [summary["peak_frame_raw"]],
            [summary["peak_energy"]],
            color="#c0504d",
            s=30,
            zorder=3,
        )
        axis.set_title(f"{case_id} 能量窗口")
        axis.set_xlabel("原始帧号")
        axis.set_ylabel("E_k (mm)")
        axis.text(
            0.02,
            0.98,
            f"ks={summary['ks_raw']}\nke={summary['ke_raw']}\n峰值帧={summary['peak_frame_raw']}",
            transform=axis.transAxes,
            va="top",
            ha="left",
            fontsize=9,
            bbox={"boxstyle": "round,pad=0.3", "facecolor": "white", "alpha": 0.9, "edgecolor": "#d0d0d0"},
        )

    handles, labels = axes[0].get_legend_handles_labels()
    figure.legend(handles, labels, loc="upper center", ncol=3, frameon=False)
    figure.tight_layout(rect=(0, 0, 1, 0.95))
    png_path = figdir / "qc_energy_windows.png"
    pdf_path = figdir / "qc_energy_windows.pdf"
    figure.savefig(png_path, dpi=600, bbox_inches="tight")
    figure.savefig(pdf_path, bbox_inches="tight")
    plt.close(figure)


def write_excel_workbook(
    workbook_path: Path,
    summary_df: pd.DataFrame,
    raw_metrics: pd.DataFrame,
    case_energy: pd.DataFrame,
    aligned_raw: pd.DataFrame,
) -> None:
    """Write a reviewer-friendly Excel workbook with core tables.

    压缩 CSV 适合后续程序读取；Excel 工作簿适合人工核对关键表格。

    Compressed CSV files are useful for downstream scripts, while the Excel
    workbook supports manual inspection of core tables.
    """
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="ks_ke_summary", index=False)
        case_energy.to_excel(writer, sheet_name="case_energy", index=False)
        raw_metrics.to_excel(writer, sheet_name="raw_metrics", index=False)
        for case_id in sorted(aligned_raw["case"].unique()):
            case_df = aligned_raw[aligned_raw["case"] == case_id].copy()
            case_df.to_excel(writer, sheet_name=case_id, index=False)


def main() -> None:
    """Run the complete displacement-alignment pipeline.

    主流程为：读取配置和位移表、计算原始指标、识别主响应窗口、生成对齐
    长表和统一网格、写出表格与质检图。

    The pipeline reads configuration and displacement tables, computes raw
    metrics, detects main response windows, writes aligned long/grid tables,
    and exports quality-control figures.
    """
    args = parse_args()
    configure_matplotlib()
    config = load_case_config(args.config)

    displacement_cfg = config["global"]["displacement"]
    grid_size = int(args.grid_size or displacement_cfg["grid_size"])
    smooth_window = int(args.smooth_window or displacement_cfg["smooth_window"])
    threshold_ratio = float(args.threshold or displacement_cfg["threshold_ratio"])

    input_dir = Path(args.input)
    outdir = Path(args.outdir)
    figdir = Path(args.figdir)
    outdir.mkdir(parents=True, exist_ok=True)
    figdir.mkdir(parents=True, exist_ok=True)

    raw_long = build_raw_long(input_dir, config)
    raw_long = apply_case_path_cropping(raw_long, config)
    raw_metrics = compute_raw_metrics(raw_long)
    case_energy = compute_case_energy(raw_long, smooth_window=smooth_window)

    # Build one response-window summary row per case before cross-case alignment.
    # 对每个工况先生成响应窗口摘要，再进行跨工况对齐。
    summary_rows = []
    for case_id in sorted(case_energy["case"].unique()):
        summary_rows.append(
            contiguous_peak_window(
                case_energy[case_energy["case"] == case_id],
                threshold_ratio=threshold_ratio,
                case_cfg=case_settings(config, case_id),
            )
        )
    summary_df = pd.DataFrame(summary_rows).sort_values("case").reset_index(drop=True)

    aligned_raw = align_raw_long(raw_long, summary_df)
    aligned_grid = interpolate_to_grid(aligned_raw, grid_size=grid_size)

    # Attach energy columns to frame-level metrics so figure scripts can reuse
    # one table instead of recomputing energy curves.
    # 将能量列并入帧级指标表，图件脚本可直接复用，避免重复计算。
    metrics_with_energy = raw_metrics.merge(
        case_energy[["case", "frame_raw", "energy_l1_mm", "energy_smooth_mm"]],
        on=["case", "frame_raw"],
        how="left",
    )

    raw_metrics_path = outdir / "raw_displacement_metrics.csv.gz"
    aligned_long_path = outdir / "aligned_displacement_long.csv.gz"
    aligned_grid_path = outdir / "aligned_displacement_grid.csv.gz"
    summary_path = outdir / "ks_ke_summary.csv"
    workbook_path = outdir / "aligned_displacement.xlsx"
    acceptance_path = outdir / "displacement_acceptance_checks.json"

    metrics_with_energy.to_csv(raw_metrics_path, index=False, compression="gzip", encoding="utf-8-sig")
    aligned_raw.to_csv(aligned_long_path, index=False, compression="gzip", encoding="utf-8-sig")
    aligned_grid.to_csv(aligned_grid_path, index=False, compression="gzip", encoding="utf-8-sig")
    summary_df.to_csv(summary_path, index=False, encoding="utf-8-sig")
    write_excel_workbook(workbook_path, summary_df, raw_metrics, case_energy, aligned_raw)
    plot_qc_energy_windows(case_energy, summary_df, figdir)

    acceptance = build_acceptance_checks(raw_metrics)
    acceptance_path.write_text(json.dumps(acceptance, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"已输出: {raw_metrics_path}")
    print(f"已输出: {aligned_long_path}")
    print(f"已输出: {aligned_grid_path}")
    print(f"已输出: {summary_path}")
    print(f"已输出: {workbook_path}")
    print(f"已输出: {acceptance_path}")


if __name__ == "__main__":
    main()
