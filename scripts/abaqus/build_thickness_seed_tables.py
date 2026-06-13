#!/usr/bin/env python3
from __future__ import annotations

"""Build seed points and audit legacy wall-thickness formulas.

本脚本读取旧壁厚工作簿，提取每个工况的缺陷点、控制点、帧号和二维
坐标。旧工作簿中的厚度值来自半径近似公式，脚本会重新计算该近似值，
并输出可供 ODB 真壁厚提取脚本使用的种子点表。

This script reads the legacy wall-thickness workbook, extracts defect/control
point metadata, frame labels, and two-dimensional coordinates for each case.
The workbook stores a radius-based approximate thickness value. The script
recomputes that approximation and exports seed points for later ODB-based
true-thickness extraction.
"""

import argparse
import math
import re
from pathlib import Path

import openpyxl
import pandas as pd

from common import ROOT, case_settings, load_case_config, nominal_inner_radius, normalize_case_id, theta_deg


# Mapping from Chinese point labels in the legacy workbook to stable English
# tokens used by the open-source CSV outputs.
# 旧工作簿中的中文点类型映射为稳定英文标记，便于后续脚本分组。
TYPE_MAP = {
    "无缺陷": "control",
    "有缺陷": "dent",
    "有缺陷（凸起）": "bulge",
}


def parse_args() -> argparse.Namespace:
    """Parse input workbook, configuration, and output directory paths.

    默认路径已经泛化为 `data/...`，开源使用者可用命令行参数替换。

    Default paths are anonymized as `data/...`; users can replace them from the
    command line.
    """
    parser = argparse.ArgumentParser(description="从旧壁厚工作簿生成种子点和公式审计表。")
    parser.add_argument(
        "--workbook",
        default=str(ROOT / "data" / "abaqus" / "wall_thickness" / "legacy_thickness.xlsx"),
        help="旧壁厚工作簿路径。",
    )
    parser.add_argument(
        "--config",
        default=str(ROOT / "scripts" / "abaqus" / "case_config.yml"),
        help="案例配置文件路径。",
    )
    parser.add_argument(
        "--outdir",
        default=str(ROOT / "output" / "spreadsheet" / "abaqus"),
        help="输出目录。",
    )
    return parser.parse_args()


def parse_frame_value(value: object) -> tuple[int, int, str]:
    """Parse workbook frame labels such as `20` or `20-2`.

    `20-2` 表示原始帧 20 中的第二个种子点。返回原始帧号、同帧顺序和
    原始文本，保证点编号可追溯。

    A label such as `20-2` means the second seed point at raw frame 20. The
    function returns the frame number, within-frame order, and original text.
    """
    text = str(value).strip()
    match = re.match(r"(\d+)(?:-(\d+))?", text)
    if not match:
        raise ValueError(f"无法解析缺陷点帧号: {value}")
    frame_raw = int(match.group(1))
    order = int(match.group(2) or 1)
    return frame_raw, order, text


def main() -> None:
    """Read the workbook, create seed rows, and write audit CSV files.

    输出 `thickness_point_seeds.csv` 用于 ODB 脚本匹配内外表面节点；
    输出 `legacy_formula_audit.csv` 用于记录旧工作簿近似厚度与脚本复算值
    是否一致。

    `thickness_point_seeds.csv` is consumed by the ODB script to match inner
    and outer surface nodes. `legacy_formula_audit.csv` records whether the
    workbook approximation matches the recomputed value.
    """
    args = parse_args()
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    config = load_case_config(args.config)
    workbook_formula = openpyxl.load_workbook(args.workbook, data_only=False)
    workbook_value = openpyxl.load_workbook(args.workbook, data_only=True)

    seed_rows: list[dict[str, object]] = []
    audit_rows: list[dict[str, object]] = []

    for sheet_name in workbook_formula.sheetnames:
        case_id = normalize_case_id(sheet_name)
        case_cfg = case_settings(config, case_id)
        reference_inner_radius = nominal_inner_radius(case_cfg)
        nominal_thickness = float(case_cfg["nominal_thickness_mm"])

        ws_formula = workbook_formula[sheet_name]
        ws_value = workbook_value[sheet_name]
        order_by_frame: dict[int, int] = {}

        for row_idx in range(2, ws_formula.max_row + 1):
            # Each non-empty row describes one seed point. The formula workbook
            # keeps the original Excel formula; the value workbook gives the
            # evaluated number.
            # 每个非空行对应一个种子点。公式版工作簿保留 Excel 原公式，
            # 数值版工作簿给出计算后的数值。
            label = ws_value.cell(row=row_idx, column=1).value
            frame_value = ws_value.cell(row=row_idx, column=2).value
            x_seed = ws_value.cell(row=row_idx, column=3).value
            y_seed = ws_value.cell(row=row_idx, column=4).value
            legacy_value = ws_value.cell(row=row_idx, column=5).value
            legacy_formula = ws_formula.cell(row=row_idx, column=5).value

            if label is None and frame_value is None:
                continue

            frame_raw, frame_order, frame_text = parse_frame_value(frame_value)
            order_by_frame[frame_raw] = max(order_by_frame.get(frame_raw, 0), frame_order)
            point_id = f"{case_id}_F{frame_raw:03d}_P{frame_order}"
            point_type = TYPE_MAP[str(label).strip()]

            x_float = float(x_seed)
            y_float = float(y_seed)
            theta_value = theta_deg(x_float, y_float)
            recomputed_legacy = math.sqrt(x_float**2 + y_float**2) - reference_inner_radius

            # The seed table keeps both geometric coordinates and legacy
            # thickness values. Later scripts can compare the approximate value
            # with ODB-derived true thickness without reopening the workbook.
            # 种子表同时保留几何坐标和旧近似厚度，后续脚本无需再打开工作簿
            # 即可比较近似厚度和 ODB 真壁厚。
            row_common = {
                "case": case_id,
                "point_id": point_id,
                "point_label": str(label).strip(),
                "point_type": point_type,
                "frame_raw": frame_raw,
                "frame_label": frame_text,
                "point_order": frame_order,
                "x_seed_mm": x_float,
                "y_seed_mm": y_float,
                "z_seed_mm": "",
                "theta_seed_deg": theta_value,
                "reference_inner_radius_mm": reference_inner_radius,
                "nominal_thickness_mm": nominal_thickness,
                "legacy_formula": legacy_formula,
                "legacy_thickness_mm": float(legacy_value),
                "legacy_thickness_recomputed_mm": recomputed_legacy,
                "legacy_bias_to_nominal_mm": float(legacy_value) - nominal_thickness,
                "seed_mode": "xy_only",
                "notes": "旧表为单点半径近似量，非内外表面成对测厚。",
            }
            seed_rows.append(row_common)
            audit_rows.append(
                {
                    **row_common,
                    "formula_matches_value": abs(float(legacy_value) - recomputed_legacy) < 1e-8,
                }
            )

    seed_df = pd.DataFrame(seed_rows).sort_values(["case", "frame_raw", "point_order", "point_id"])
    audit_df = pd.DataFrame(audit_rows).sort_values(["case", "frame_raw", "point_order", "point_id"])

    seed_path = outdir / "thickness_point_seeds.csv"
    audit_path = outdir / "legacy_formula_audit.csv"
    seed_df.to_csv(seed_path, index=False, encoding="utf-8-sig")
    audit_df.to_csv(audit_path, index=False, encoding="utf-8-sig")

    print(f"已输出种子点表: {seed_path}")
    print(f"已输出公式审计表: {audit_path}")


if __name__ == "__main__":
    main()
